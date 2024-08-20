from .db_connection import get_mongo_client
from celery import shared_task
import csv
import io
from openpyxl import load_workbook
from django.conf import settings
import pandas as pd
from pathlib import Path
from datetime import datetime
from celery.exceptions import SoftTimeLimitExceeded

@shared_task(bind=True, max_retries=5)
def process_excel_file(self, file_path, start_row=2):
    client = None
    wb = None
    try:
        # Connect to MongoDB
        client = get_mongo_client()
        db = client[settings.DB_NAME]

        # Determine collection name from file name
        collection_name = Path(file_path).name.split(".")[0]
        collection = db[collection_name]

        # Load workbook
        wb = load_workbook(filename=file_path, read_only=True)
        sheet = wb.active

        # Get headers
        headers = [cell.value if cell.value is not None else f"unnamed_column_{i}" 
                   for i, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)))]

        schema = {header: set() for header in headers}
        chunk_size = 1000
        records = []
        processed_rows = 0

        for row_index, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row):
            record = {}
            for header, cell in zip(headers, row):
                value = cell.value

                if value is None:
                    record[header] = None
                else:
                    value_type = get_value_type(value)
                    schema[header].add(value_type)

                    if value_type == 'date':
                        value = value.isoformat()

                record[header] = value

            records.append(record)
            processed_rows += 1

            if len(records) >= chunk_size:
                collection.insert_many(records)
                records = []
                # Update task state
                self.update_state(state='PROGRESS', meta={'current': row_index, 'total': sheet.max_row})

        # Insert any remaining records
        if records:
            collection.insert_many(records)

        # Finalize and store schema
        final_schema = finalize_schema(schema)
        store_schema(db, collection_name, final_schema)

        print(f"Processed {file_path} into {collection_name} collection. Total rows: {processed_rows}")

    except SoftTimeLimitExceeded:
        # If we hit a time limit, save progress and retry
        if records:
            collection.insert_many(records)
        next_row = start_row + processed_rows
        raise self.retry(args=(file_path, next_row))

    except Exception as e:
        # For other exceptions, retry with the next unprocessed row
        next_row = start_row + processed_rows
        raise self.retry(args=(file_path, next_row), exc=e)

    finally:
        if client:
            client.close()
        if wb:
            wb.close()

def get_value_type(value):
    if value is None:
        return 'null'
    elif isinstance(value, str):
        return 'string'
    elif isinstance(value, int):
        return 'int'
    elif isinstance(value, float):
        return 'float'
    elif isinstance(value, bool):
        return 'bool'
    elif isinstance(value, datetime):
        return 'date'
    else:
        return 'string'

def finalize_schema(schema):
    final_schema = {}
    for header, types in schema.items():
        if 'null' in types:
            types.remove('null')
        if len(types) == 1:
            final_schema[header] = types.pop()
        else:
            final_schema[header] = 'string'
    return final_schema

def store_schema(db, collection_name, schema):
    schema_collection = db['schemas']
    schema_collection.update_one(
        {'collection': collection_name},
        {'$set': {'schema': schema}},
        upsert=True
    )

@shared_task
def excel_to_csv(excel_file):
    # Load the Excel file
    wb = load_workbook(excel_file.file.path)
    sheet = wb.active

    # Create a CSV file in memory
    csv_buffer = io.StringIO()
    csv_writer = csv.writer(csv_buffer)

    # Write data from Excel to CSV
    for row in sheet.iter_rows(values_only=True):
        csv_writer.writerow(row)
    
    return csv_buffer