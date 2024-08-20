from django.shortcuts import render
from django.core.paginator import Paginator
from rest_framework.response import Response
from rest_framework.decorators import api_view
from .models import ExcelFile
from openpyxl import load_workbook
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .utils import generate_md5_hash, file_exists
from .tasks import excel_to_csv, process_excel_file

@api_view(['POST'])
def upload_file(request):
    file = request.FILES['file']
    hash = generate_md5_hash(file)

    if file_exists(hash):
        return Response("File already exists!")

    excel_file = ExcelFile.objects.create(
        file=file,
        name=file.name,
        size=file.size,
        hash=hash
    )
    task = process_excel_file.delay(excel_file.file.path)
    return Response({'id': excel_file.id, 'name': excel_file.name, "task_id": str(task.id)})

@api_view(['GET'])
def list_files(request):
    files = ExcelFile.objects.all().values('id', 'name', 'size', 'uploaded_at')
    return Response(list(files))

@api_view(['GET'])
def get_columns(request, file_id):
    excel_file = ExcelFile.objects.get(id=file_id)
    
    # Load only the worksheet properties and header row
    wb = load_workbook(filename=excel_file.file.path, read_only=True)
    ws = wb.active
    
    # Get the first row (header)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    
    # Close the workbook to free up resources
    wb.close()
    
    # Convert header row to list and filter out None values
    columns = [col for col in header_row if col is not None]
    
    return Response({'columns': columns})

@api_view(['POST'])
def get_data(request):
    excel_file = ExcelFile.objects.get(id=request.data.get('id'))
    selected_columns = request.data.get('columns')
    page_number = int(request.data.get('page', 1))
    items_per_page = int(request.data.get('rowsPerPage', 10))

    wb = load_workbook(filename=excel_file.file.path, read_only=True)
    ws = wb.active

    # Get the indices of selected columns
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_indices = [header_row.index(col) for col in selected_columns if col in header_row]

    # Calculate the range of rows to read
    start_row = (page_number - 1) * items_per_page + 2  # +2 to skip header and account for 1-indexing
    end_row = start_row + items_per_page

    records = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        record = {selected_columns[i]: row[col_indices[i]] for i in range(len(selected_columns))}
        records.append(record)

    total_rows = ws.max_row - 1  # Subtract 1 to exclude header row
    total_pages = (total_rows + items_per_page - 1) // items_per_page

    wb.close()

    return Response({
        'records': records,
        'page': page_number,
        'totalPages': total_pages,
        'totalRecords': total_rows,
    })

@api_view(['GET'])
def convert_excel_to_csv(request, pk):
    # Get the ExcelFile instance
    excel_file = get_object_or_404(ExcelFile, pk=pk)

    if not excel_file.file:
        return Response({'error': 'No file associated with this record'}, status=status.HTTP_400_BAD_REQUEST)
    
    csv_buffer = excel_to_csv(excel_file)
    # Prepare the response
    response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{excel_file.name}.csv"'
    return response
